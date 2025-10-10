import os
import json
from datetime import datetime
from typing import Optional, Dict, Any, List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging
import pandas as pd

logger = logging.getLogger(__name__)

ORDER_LOG_FILE = "orders.xlsx"

class OrderLogger:
    def __init__(self, file_path: str = ORDER_LOG_FILE):
        self.file_path = file_path
        self.ensure_order_file_exists()
    
    def ensure_order_file_exists(self):
        """Create the order log Excel file if it doesn't exist."""
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Orders"
            
            # Add headers with styling
            headers = [
                "order_id", "timestamp", "customer_phone", "customer_name", 
                "order_type", "total_amount", "currency", "status", 
                "catalog_id", "order_text", "products_json", "admin_notes",
                "processed_by", "processing_timestamp", "delivery_address", "payment_method"
            ]
            ws.append(headers)

    def search_orders(self, query: str, criteria: str = 'all') -> List[Dict[str, Any]]:
        """
        Search orders based on various criteria.
        
        Args:
            query (str): The search query
            criteria (str): The search criteria ('phone', 'order_id', 'name', 'status', 'all')
        
        Returns:
            List of matching orders
        """
        try:
            df = pd.read_excel(self.file_path)
            query = str(query).lower()
            
            if criteria == 'phone':
                mask = df['customer_phone'].astype(str).str.contains(query, case=False, na=False)
            elif criteria == 'order_id':
                mask = df['order_id'].astype(str).str.contains(query, case=False, na=False)
            elif criteria == 'name':
                mask = df['customer_name'].astype(str).str.contains(query, case=False, na=False)
            elif criteria == 'status':
                mask = df['status'].astype(str).str.contains(query, case=False, na=False)
            else:  # 'all'
                mask = df.apply(lambda x: x.astype(str).str.contains(query, case=False, na=False)).any(axis=1)
            
            results = df[mask].to_dict('records')
            return results
        except Exception as e:
            logger.exception(f"Failed to search orders: {e}")
            return []
            
            # Style the header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                # Auto-adjust column width
                ws.column_dimensions[cell.column_letter].width = max(len(header) + 2, 12)
            
            wb.save(self.file_path)
            logger.info(f"Created order log file: {self.file_path}")
    
    def log_order(
        self,
        customer_phone: str,
        customer_name: str,
        order_type: str,
        total_amount: float,
        catalog_id: str,
        order_text: str,
        products_data: List[Dict[str, Any]],
        currency: str = "USD",
        status: str = "NEW"
    ) -> str:
        """Log an order to the Excel file and return the generated order ID."""
        try:
            # Generate unique order ID
            timestamp = datetime.now()
            order_id = f"ORD{timestamp.strftime('%Y%m%d%H%M%S')}{customer_phone[-4:]}"
            
            # Load existing workbook
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Prepare data
            timestamp_str = timestamp.strftime("%Y-%m-%d %H:%M:%S")
            products_json = json.dumps(products_data)
            
            # Add row
            row_data = [
                order_id,
                timestamp_str,
                customer_phone,
                customer_name,
                order_type,
                total_amount,
                currency,
                status,
                catalog_id,
                order_text,
                products_json,
                "",  # admin_notes
                "",  # processed_by
                "",  # processing_timestamp
                "",  # delivery_address
                ""   # payment_method
            ]
            
            row_num = ws.max_row + 1
            ws.append(row_data)
            
            # Style the new row
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Color coding by status
                if col_num == 8:  # status column
                    if status == "NEW":
                        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    elif status == "PROCESSING":
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    elif status == "COMPLETED":
                        cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            
            wb.save(self.file_path)
            logger.info(f"Logged order: {order_id} for {customer_phone}")
            return order_id
            
        except Exception as e:
            logger.exception(f"Failed to log order: {e}")
            return f"ORD_ERROR_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    def export_orders(self, criteria: dict = None) -> str:
        """
        Export orders to a new Excel file with optional filtering.
        
        Args:
            criteria (dict): Filter criteria (e.g., {'status': 'COMPLETED', 'date': '2025-10'})
        
        Returns:
            str: Path to the exported file
        """
        try:
            df = pd.read_excel(self.file_path)
            
            if criteria:
                if 'status' in criteria:
                    df = df[df['status'] == criteria['status']]
                if 'date' in criteria:
                    df = df[df['timestamp'].str.contains(criteria['date'])]
                if 'customer' in criteria:
                    mask = (df['customer_name'].str.contains(criteria['customer'], case=False, na=False) |
                           df['customer_phone'].str.contains(criteria['customer'], case=False, na=False))
                    df = df[mask]
            
            # Generate export filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_path = f"orders_export_{timestamp}.xlsx"
            
            # Export with styling
            wb = Workbook()
            ws = wb.active
            ws.title = "Orders Export"
            
            # Write headers
            headers = list(df.columns)
            ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # Write data
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == headers.index('status') + 1:  # Status column
                        status = str(value).upper()
                        if status == "COMPLETED":
                            cell.fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                        elif status == "CANCELLED":
                            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            wb.save(export_path)
            return export_path
            
        except Exception as e:
            logger.exception(f"Failed to export orders: {e}")
            return ""

    def update_order_status(self, order_id: str, status: str, admin_notes: str = "", processed_by: str = "") -> bool:
        """Update order status and add admin notes."""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Find the order row
            for row_num in range(2, ws.max_row + 1):  # Skip header
                if ws.cell(row=row_num, column=1).value == order_id:
                    # Update status
                    ws.cell(row=row_num, column=8).value = status
                    
                    # Update admin notes
                    if admin_notes:
                        current_notes = ws.cell(row=row_num, column=12).value or ""
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                        new_note = f"[{timestamp}] {admin_notes}"
                        updated_notes = f"{current_notes}\n{new_note}" if current_notes else new_note
                        ws.cell(row=row_num, column=12).value = updated_notes
                    
                    # Update processed by and timestamp
                    if processed_by:
                        ws.cell(row=row_num, column=13).value = processed_by
                        ws.cell(row=row_num, column=14).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # Update cell styling based on status
                    status_cell = ws.cell(row=row_num, column=8)
                    if status == "NEW":
                        status_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    elif status == "PROCESSING":
                        status_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    elif status == "COMPLETED":
                        status_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    elif status == "CANCELLED":
                        status_cell.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                    
                    wb.save(self.file_path)
                    logger.info(f"Updated order {order_id} status to {status}")
                    return True
            
            logger.warning(f"Order {order_id} not found for status update")
            return False
            
        except Exception as e:
            logger.exception(f"Failed to update order status: {e}")
            return False
    
    def get_orders_by_status(self, status: str = None) -> List[Dict[str, Any]]:
        """Get orders filtered by status."""
        try:
            if not os.path.exists(self.file_path):
                return []
            
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            orders = []
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # order_id
                    continue
                
                order_data = dict(zip(headers, row))
                
                # Filter by status if specified
                if status and order_data.get('status') != status:
                    continue
                
                orders.append(order_data)
            
            return orders
            
        except Exception as e:
            logger.exception(f"Failed to get orders by status: {e}")
            return []
    
    def get_recent_orders(self, limit: int = 10) -> List[Dict[str, Any]]:
        """Get recent orders."""
        try:
            if not os.path.exists(self.file_path):
                return []
            
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            orders = []
            headers = [cell.value for cell in ws[1]]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            # Get last N rows (most recent)
            for row in rows[-limit:]:
                if row[0]:  # order_id exists
                    orders.append(dict(zip(headers, row)))
            
            return list(reversed(orders))  # Most recent first
            
        except Exception as e:
            logger.exception(f"Failed to get recent orders: {e}")
            return []
    
    def get_order_details(self, order_id: str) -> Optional[Dict[str, Any]]:
        """Get detailed information about a specific order."""
        try:
            if not os.path.exists(self.file_path):
                return None
            
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == order_id:
                    order_data = dict(zip(headers, row))
                    
                    # Parse products JSON
                    try:
                        if order_data.get('products_json'):
                            order_data['products'] = json.loads(order_data['products_json'])
                        else:
                            order_data['products'] = []
                    except:
                        order_data['products'] = []
                    
                    return order_data
            
            return None
            
        except Exception as e:
            logger.exception(f"Failed to get order details: {e}")
            return None
    
    def get_order_statistics(self) -> Dict[str, Any]:
        """Get order statistics for admin dashboard."""
        try:
            if not os.path.exists(self.file_path):
                return {
                    "total_orders": 0,
                    "new_orders": 0,
                    "processing_orders": 0,
                    "completed_orders": 0,
                    "cancelled_orders": 0,
                    "total_revenue": 0,
                    "average_order_value": 0
                }
            
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            stats = {
                "total_orders": 0,
                "new_orders": 0,
                "processing_orders": 0,
                "completed_orders": 0,
                "cancelled_orders": 0,
                "total_revenue": 0,
                "average_order_value": 0
            }
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # order_id
                    continue
                
                stats["total_orders"] += 1
                
                # Count by status
                status = row[7] or "NEW"
                if status == "NEW":
                    stats["new_orders"] += 1
                elif status == "PROCESSING":
                    stats["processing_orders"] += 1
                elif status == "COMPLETED":
                    stats["completed_orders"] += 1
                elif status == "CANCELLED":
                    stats["cancelled_orders"] += 1
                
                # Calculate revenue (only completed orders)
                if status == "COMPLETED":
                    try:
                        amount = float(row[5] or 0)
                        stats["total_revenue"] += amount
                    except:
                        pass
            
            # Calculate average order value
            if stats["completed_orders"] > 0:
                stats["average_order_value"] = round(stats["total_revenue"] / stats["completed_orders"], 2)
            
            return stats
            
        except Exception as e:
            logger.exception(f"Failed to get order statistics: {e}")
            return {"error": str(e)}

# Global order logger instance
order_logger = OrderLogger()
