import os
import logging
import inspect
from typing import List, Tuple
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

LAPTOPS_EXCEL_FILE = "laptops.xlsx"
REPAIRS_EXCEL_FILE = "repairs.xlsx"


def _read_ids_from_excel(filepath: str) -> List[str]:
    """Read retailer IDs from the first column of an Excel file."""
    ids: List[str] = []
    if not os.path.exists(filepath):
        logger.info("Excel file not found at %s", filepath)
        return ids
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active  # Use the active sheet
        
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            val = row[0]
            if not val:
                continue
            s = str(val).strip()
            if not s or s.lower() in ("retailer_id", "(none configured)"):
                continue
            ids.append(s)
    except Exception as exc:
        logger.exception("Failed to open Excel file %s: %s", filepath, exc)
    
    return ids


def load_laptop_retailer_ids() -> List[str]:
    """Load laptop retailer IDs from laptops.xlsx."""
    return _read_ids_from_excel(LAPTOPS_EXCEL_FILE)


def load_repair_retailer_ids() -> List[str]:
    """Load repair retailer IDs from repairs.xlsx."""
    return _read_ids_from_excel(REPAIRS_EXCEL_FILE)


def load_retailer_ids_from_excel(filepath: str = None) -> Tuple[List[str], List[str]]:
    """Legacy function for backward compatibility - loads from both files."""
    laptop_ids = load_laptop_retailer_ids()
    repair_ids = load_repair_retailer_ids()
    return laptop_ids, repair_ids


def env_retailer_ids(*keys: str) -> List[str]:
    ids = []
    for k in keys:
        v = os.getenv(k)
        if v and v.strip():
            ids.append(v.strip())
    return ids


def send_catalog_compat(whatsapp, to: str, retailer_ids: List[str], header: str, body: str, footer: str, catalog_id: str = None, fallback_button_id: str = "browse_catalog"):
    """Send a catalog/product list using send_catalog_product_list with CatalogSection.

    Falls back to an interactive-button message or plain text listing of retailer ids.
    """
    try:
        # Use send_catalog_product_list with CatalogSection for specific retailer IDs
        if hasattr(whatsapp, "send_catalog_product_list") and catalog_id:
            from wa_cloud_py.components.messages import CatalogSection
            
            # Create a single catalog section with the specific retailer IDs
            product_sections = [CatalogSection(
                title="Products",
                retailer_product_ids=retailer_ids  # Correct parameter name
            )]
            
            logger.info("Sending catalog with product sections containing retailer IDs: %s", retailer_ids)
            return whatsapp.send_catalog_product_list(
                to=to,
                catalog_id=catalog_id,
                header=header,
                body=body,
                product_sections=product_sections,
                footer=footer,
            )

        # Fallback: try other methods if send_catalog_product_list is not available
        if hasattr(whatsapp, "send_product_list"):
            return whatsapp.send_product_list(
                to=to,
                retailer_ids=retailer_ids,
                header=header,
                body=body,
                footer=footer,
            )

    except Exception as exc:
        logger.exception("Failed to send catalog with specific retailer IDs: %s", exc)

    # Final fallback: interactive list showing the specific retailer IDs
    lines = [f"- {rid}" for rid in retailer_ids]
    body_with_ids = f"{body}\n\nConfigured retailer IDs:\n" + "\n".join(lines)
    try:
        return whatsapp.send_interactive_buttons(
            to=to,
            body=body_with_ids,
            buttons=[__import__('wa_cloud_py').components.messages.ReplyButton(id=fallback_button_id, title="Browse Catalog")],
        )
    except Exception:
        try:
            return whatsapp.send_text(to=to, body=body_with_ids)
        except Exception:
            logger.exception("Failed to send fallback catalog message")
            return None
