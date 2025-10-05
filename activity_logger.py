import os
import json
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logger = logging.getLogger(__name__)

ACTIVITY_LOG_FILE = "activity_log.xlsx"

class ActivityLogger:
    def __init__(self, file_path: str = ACTIVITY_LOG_FILE):
        self.file_path = file_path
        self.ensure_log_file_exists()
    
    def ensure_log_file_exists(self):
        """Create the activity log Excel file if it doesn't exist."""
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Activity_Log"
            
            # Add headers with styling
            headers = [
                "timestamp", "phone_number", "user_name", "activity_type", 
                "message_type", "user_input", "bot_response", "button_id",
                "admin_flag", "session_id", "additional_data"
            ]
            ws.append(headers)
            
            # Style the header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            wb.save(self.file_path)
            logger.info(f"Created activity log file: {self.file_path}")
    
    def log_activity(
        self,
        phone_number: str,
        activity_type: str,
        user_name: str = None,
        message_type: str = None,
        user_input: str = None,
        bot_response: str = None,
        button_id: str = None,
        admin_flag: bool = False,
        session_id: str = None,
        additional_data: Dict[str, Any] = None
    ):
        """Log an activity to the Excel file."""
        try:
            # Load existing workbook
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Prepare data
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            additional_data_json = json.dumps(additional_data) if additional_data else None
            
            # Truncate long text fields for Excel compatibility
            user_input = (user_input[:500] + "...") if user_input and len(user_input) > 500 else user_input
            bot_response = (bot_response[:500] + "...") if bot_response and len(bot_response) > 500 else bot_response
            
            # Add row
            row_data = [
                timestamp,
                phone_number,
                user_name,
                activity_type,
                message_type,
                user_input,
                bot_response,
                button_id,
                admin_flag,
                session_id,
                additional_data_json
            ]
            
            ws.append(row_data)
            wb.save(self.file_path)
            
            logger.info(f"Logged activity: {activity_type} for {phone_number}")
            
        except Exception as e:
            logger.exception(f"Failed to log activity: {e}")
    
    def get_analytics_summary(self, days: int = 7) -> Dict[str, Any]:
        """Get comprehensive analytics summary for the last N days."""
        try:
            if not os.path.exists(self.file_path):
                return {"error": "No activity data found"}
            
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            # Calculate date threshold
            threshold_date = datetime.now() - timedelta(days=days)
            
            activities = []
            unique_users = set()
            admin_activities = 0
            total_conversations = 0
            activity_types = {}
            hourly_activity = {str(i): 0 for i in range(24)}
            daily_activity = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # timestamp
                    continue
                
                try:
                    timestamp_str = str(row[0])
                    if len(timestamp_str) > 19:  # Handle datetime objects
                        timestamp_str = timestamp_str[:19]
                    
                    activity_time = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                    
                    if activity_time >= threshold_date:
                        activities.append(row)
                        unique_users.add(row[1])  # phone_number
                        
                        if row[8]:  # admin_flag
                            admin_activities += 1
                        
                        # Activity type counting
                        activity_type = row[3] or "unknown"
                        activity_types[activity_type] = activity_types.get(activity_type, 0) + 1
                        
                        # Hourly distribution
                        hour = str(activity_time.hour)
                        hourly_activity[hour] += 1
                        
                        # Daily distribution
                        day = activity_time.strftime("%Y-%m-%d")
                        daily_activity[day] = daily_activity.get(day, 0) + 1
                        
                except Exception as e:
                    logger.warning(f"Error parsing activity row: {e}")
                    continue
            
            # Calculate conversation metrics
            session_ids = set(row[9] for row in activities if row[9])
            avg_activities_per_user = len(activities) / len(unique_users) if unique_users else 0
            
            # Top activity types
            top_activities = sorted(activity_types.items(), key=lambda x: x[1], reverse=True)[:5]
            
            # Peak hours
            peak_hours = sorted(hourly_activity.items(), key=lambda x: x[1], reverse=True)[:3]
            
            return {
                "period_days": days,
                "total_activities": len(activities),
                "unique_users": len(unique_users),
                "admin_activities": admin_activities,
                "user_activities": len(activities) - admin_activities,
                "total_sessions": len(session_ids),
                "avg_activities_per_user": round(avg_activities_per_user, 2),
                "top_activity_types": top_activities,
                "peak_hours": peak_hours,
                "daily_breakdown": daily_activity,
                "hourly_breakdown": hourly_activity
            }
            
        except Exception as e:
            logger.exception(f"Failed to get analytics summary: {e}")
            return {"error": str(e)}
    
    def get_conversation_analytics(self, phone_number: str = None) -> Dict[str, Any]:
        """Get detailed conversation analytics for a specific user or all users."""
        try:
            if not os.path.exists(self.file_path):
                return {"error": "No activity data found"}
            
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            conversations = {}
            user_stats = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # timestamp
                    continue
                
                user_phone = row[1]
                session_id = row[9]
                timestamp_str = str(row[0])[:19]
                
                # Filter by phone number if specified
                if phone_number and user_phone != phone_number:
                    continue
                
                try:
                    activity_time = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                    
                    # Track by session
                    if session_id:
                        if session_id not in conversations:
                            conversations[session_id] = {
                                "user": user_phone,
                                "start_time": activity_time,
                                "end_time": activity_time,
                                "activity_count": 0,
                                "activities": []
                            }
                        
                        conv = conversations[session_id]
                        conv["end_time"] = max(conv["end_time"], activity_time)
                        conv["activity_count"] += 1
                        conv["activities"].append(row[3])  # activity_type
                    
                    # Track by user
                    if user_phone not in user_stats:
                        user_stats[user_phone] = {
                            "total_activities": 0,
                            "sessions": set(),
                            "first_activity": activity_time,
                            "last_activity": activity_time,
                            "activity_types": {}
                        }
                    
                    user_stat = user_stats[user_phone]
                    user_stat["total_activities"] += 1
                    user_stat["last_activity"] = max(user_stat["last_activity"], activity_time)
                    if session_id:
                        user_stat["sessions"].add(session_id)
                    
                    activity_type = row[3] or "unknown"
                    user_stat["activity_types"][activity_type] = user_stat["activity_types"].get(activity_type, 0) + 1
                    
                except Exception as e:
                    logger.warning(f"Error parsing conversation row: {e}")
                    continue
            
            # Calculate conversation durations
            conversation_durations = []
            for conv in conversations.values():
                duration = (conv["end_time"] - conv["start_time"]).total_seconds() / 60  # minutes
                conversation_durations.append(duration)
            
            # Calculate user engagement metrics
            engagement_metrics = {}
            for phone, stats in user_stats.items():
                total_time = (stats["last_activity"] - stats["first_activity"]).total_seconds() / 60
                engagement_metrics[phone] = {
                    "total_activities": stats["total_activities"],
                    "session_count": len(stats["sessions"]),
                    "total_engagement_minutes": round(total_time, 2),
                    "avg_activities_per_session": round(stats["total_activities"] / max(len(stats["sessions"]), 1), 2),
                    "top_activity": max(stats["activity_types"].items(), key=lambda x: x[1])[0] if stats["activity_types"] else "none"
                }
            
            avg_duration = sum(conversation_durations) / len(conversation_durations) if conversation_durations else 0
            
            return {
                "total_conversations": len(conversations),
                "total_users": len(user_stats),
                "avg_conversation_duration_minutes": round(avg_duration, 2),
                "conversation_durations": conversation_durations,
                "user_engagement": engagement_metrics,
                "longest_conversation_minutes": max(conversation_durations) if conversation_durations else 0,
                "shortest_conversation_minutes": min(conversation_durations) if conversation_durations else 0
            }
            
        except Exception as e:
            logger.exception(f"Failed to get conversation analytics: {e}")
            return {"error": str(e)}
    
    def export_filtered_data(self, 
                           start_date: str = None, 
                           end_date: str = None, 
                           activity_types: List[str] = None,
                           admin_only: bool = False,
                           output_file: str = "filtered_activity_export.xlsx") -> bool:
        """Export filtered activity data to a new Excel file."""
        try:
            if not os.path.exists(self.file_path):
                return False
            
            wb_source = load_workbook(self.file_path, read_only=True)
            ws_source = wb_source.active
            
            # Create new workbook for export
            wb_export = Workbook()
            ws_export = wb_export.active
            ws_export.title = "Filtered_Activity_Log"
            
            # Copy headers
            headers = [cell.value for cell in ws_source[1]]
            ws_export.append(headers)
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws_export.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Parse date filters
            start_dt = None
            end_dt = None
            if start_date:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            if end_date:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)  # Include end date
            
            # Filter and copy data
            exported_rows = 0
            for row in ws_source.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # timestamp
                    continue
                
                try:
                    # Date filtering
                    timestamp_str = str(row[0])[:19]
                    activity_time = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                    
                    if start_dt and activity_time < start_dt:
                        continue
                    if end_dt and activity_time >= end_dt:
                        continue
                    
                    # Activity type filtering
                    if activity_types and row[3] not in activity_types:
                        continue
                    
                    # Admin filtering
                    if admin_only and not row[8]:  # admin_flag
                        continue
                    
                    ws_export.append(row)
                    exported_rows += 1
                    
                except Exception as e:
                    logger.warning(f"Error filtering row: {e}")
                    continue
            
            # Add summary sheet
            ws_summary = wb_export.create_sheet("Export_Summary")
            summary_data = [
                ["Export Summary", ""],
                ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Total Rows Exported", exported_rows],
                ["Date Range", f"{start_date or 'All'} to {end_date or 'All'}"],
                ["Activity Types", ", ".join(activity_types) if activity_types else "All"],
                ["Admin Only", "Yes" if admin_only else "No"],
            ]
            
            for row_data in summary_data:
                ws_summary.append(row_data)
            
            wb_export.save(output_file)
            logger.info(f"Exported {exported_rows} rows to {output_file}")
            return True
            
        except Exception as e:
            logger.exception(f"Failed to export filtered data: {e}")
            return False
        """Get total activity count for a user."""
        try:
            if not os.path.exists(self.file_path):
                return 0
                
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[1] == phone_number:  # phone_number column
                    count += 1
            
            return count
        except Exception as e:
            logger.exception(f"Failed to get activity count: {e}")
            return 0
    
    def get_user_activity_count(self, phone_number: str) -> int:
        """Get total activity count for a user."""
        try:
            if not os.path.exists(self.file_path):
                return 0
                
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[1] == phone_number:  # phone_number column
                    count += 1
            
            return count
        except Exception as e:
            logger.exception(f"Failed to get activity count: {e}")
            return 0
        """Get recent activities for admin dashboard."""
        try:
            if not os.path.exists(self.file_path):
                return []
                
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            activities = []
            rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
            
            # Get last N rows (most recent)
            for row in rows[-limit:]:
                if row[0]:  # timestamp exists
                    activities.append({
                        'timestamp': row[0],
                        'phone_number': row[1],
                        'user_name': row[2],
                        'activity_type': row[3],
                        'admin_flag': row[8]
                    })
            
            return list(reversed(activities))  # Most recent first
        except Exception as e:
            logger.exception(f"Failed to get recent activities: {e}")
            return []

# Global activity logger instance
activity_logger = ActivityLogger()
