import os
from dotenv import load_dotenv
from fastapi import FastAPI
from fastapi.staticfiles import StaticFiles
from wa_cloud_py import WhatsApp
from fastapi.responses import PlainTextResponse
from fastapi import Request, HTTPException
from wa_cloud_py.components.messages import (
    CatalogSection,
    ImageHeader,
    ListSection,
    ReplyButton,
)
from wa_cloud_py.messages.types import InteractiveButtonMessage, OrderMessage

from wa_cloud_py.messages.types import (
    TextMessage,
)
import logging
import uvicorn
import asyncio
import inspect
import uuid
from datetime import datetime, timedelta
from typing import List, Tuple, Optional
from openpyxl import load_workbook
from activity_logger import activity_logger
from order_logger import order_logger


load_dotenv()

VERIFY_TOKEN = os.getenv("VERIFY_TOKEN")
ACCESS_TOKEN = os.getenv("ACCESS_TOKEN")
PHONE_NUMBER_ID = os.getenv("PHONE_NUMBER_ID")
CATALOG_ID = os.getenv("CATALOG_ID")
PRODUCT_RETAILER_ID = os.getenv("PRODUCT_RETAILER_ID")
PRODUCT_RETAILER_ID_2 = os.getenv("PRODUCT_RETAILER_ID_2")
PRODUCT_RETAILER_ID_REPAIR = os.getenv("PRODUCT_RETAILER_ID_REPAIR")
PRODUCT_RETAILER_ID_REPAIR_2 = os.getenv("PRODUCT_RETAILER_ID_REPAIR_2")
PUBLIC_URL = os.getenv("PUBLIC_URL")

# Admin configuration
ADMIN_NUMBER = "263711475883"

if not VERIFY_TOKEN:
    raise ValueError("VERIFY_TOKEN environment variable is not set")
if not ACCESS_TOKEN:
    raise ValueError("ACCESS_TOKEN environment variable is not set")
if not PHONE_NUMBER_ID:
    raise ValueError("PHONE_NUMBER_ID environment variable is not set")
if not CATALOG_ID:
    raise ValueError("CATALOG_ID environment variable is not set")
if not PRODUCT_RETAILER_ID:
    raise ValueError("PRODUCT_RETAILER_ID environment variable is not set")
# PRODUCT_RETAILER_ID_2 is optional — no raise
if not PUBLIC_URL:
    raise ValueError("PUBLIC_URL environment variable is not set")


app = FastAPI()

# Mount static files to serve the video
app.mount("/static", StaticFiles(directory="."), name="static")


whatsapp = WhatsApp(access_token=ACCESS_TOKEN, phone_number_id=PHONE_NUMBER_ID)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EXCEL_RETAILER_FILE = "spectrax_retailer_ids.xlsx"


def generate_session_id() -> str:
    """Generate a unique session ID for tracking conversations."""
    return str(uuid.uuid4())[:8]


def _read_ids_from_sheet(workbook, sheet_name: str) -> List[str]:
    """Read first-column values from a sheet, skipping header and placeholders."""
    ids: List[str] = []
    if sheet_name not in workbook.sheetnames:
        return ids
    ws = workbook[sheet_name]
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        if row[0] and row[0] != "retailer_id":  # Skip header and empty rows
            ids.append(str(row[0]))
    return ids


def load_retailer_ids_from_excel(filepath: str = EXCEL_RETAILER_FILE) -> Tuple[List[str], List[str]]:
    """Return (laptop_ids, repair_ids). If file not found or empty, return empty lists."""
    if not os.path.exists(filepath):
        return [], []
    try:
        wb = load_workbook(filepath, read_only=True)
    except Exception as exc:
        logger.exception("Failed to load Excel file: %s", exc)
        return [], []
    laptop_ids = _read_ids_from_sheet(wb, "Laptops")
    repair_ids = _read_ids_from_sheet(wb, "Repairs")
    return laptop_ids, repair_ids


def _env_retailer_ids(*keys: str) -> List[str]:
    ids = []
    for k in keys:
        val = os.getenv(k)
        if val:
            ids.append(val)
    return ids


def safe_mark_as_read(message_id: str):
    """Safely mark a WhatsApp message as read; swallow and log errors from the API.

    Some incoming webhook messages reference message IDs that cannot be marked as
    read (for example, legacy or unsupported IDs). Calling the API with an invalid
    ID raises an OAuthException; we log it and continue so the webhook stays healthy.
    """
    try:
        whatsapp.mark_as_read(message_id)
    except Exception as exc:
        logger.warning("Failed to mark message as read: %s", exc)


def is_admin(phone_number: str) -> bool:
    """Check if the phone number is an admin."""
    return phone_number == ADMIN_NUMBER


def handle_admin_command(phone_number: str, message_text: str):
    """Handle admin commands for managing retailer IDs."""
    if not is_admin(phone_number):
        return False
    
    message_lower = message_text.lower().strip()
    
    # Admin help command
    if message_lower in ["/admin", "/help", "help"]:
        send_admin_help(phone_number)
        return True
    
    # Add laptop retailer ID
    if message_lower.startswith("/add_laptop "):
        retailer_id = message_text[12:].strip()
        if retailer_id:
            add_laptop_retailer_id(phone_number, retailer_id)
        else:
            whatsapp.send_text(to=phone_number, body="❌ Please provide a retailer ID. Format: /add_laptop <retailer_id>")
        return True
    
    # Add repair retailer ID
    if message_lower.startswith("/add_repair "):
        retailer_id = message_text[12:].strip()
        if retailer_id:
            add_repair_retailer_id(phone_number, retailer_id)
        else:
            whatsapp.send_text(to=phone_number, body="❌ Please provide a retailer ID. Format: /add_repair <retailer_id>")
        return True
    
    # List current IDs
    if message_lower in ["/list", "/list_ids"]:
        list_current_retailer_ids(phone_number)
        return True
    
    # Remove laptop retailer ID
    if message_lower.startswith("/remove_laptop "):
        retailer_id = message_text[15:].strip()
        if retailer_id:
            remove_laptop_retailer_id(phone_number, retailer_id)
        else:
            whatsapp.send_text(to=phone_number, body="❌ Please provide a retailer ID. Format: /remove_laptop <retailer_id>")
        return True
    
    # Remove repair retailer ID
    if message_lower.startswith("/remove_repair "):
        retailer_id = message_text[15:].strip()
        if retailer_id:
            remove_repair_retailer_id(phone_number, retailer_id)
        else:
            whatsapp.send_text(to=phone_number, body="❌ Please provide a retailer ID. Format: /remove_repair <retailer_id>")
        return True
    
    # Order management commands
    if message_lower.startswith("/update_order "):
        return handle_admin_order_update(phone_number, message_text)
    
    # Order search command
    if message_lower.startswith("/order "):
        order_id = message_text[7:].strip().upper()
        if order_id:
            order_details = order_logger.get_order_details(order_id)
            if order_details:
                send_order_details_message(phone_number, order_details)
            else:
                whatsapp.send_text(to=phone_number, body=f"❌ Order {order_id} not found.")
        else:
            whatsapp.send_text(to=phone_number, body="❌ Please provide an order ID. Format: /order <ORDER_ID>")
        return True
    
    return False


def send_order_details_message(phone_number: str, order_details: dict):
    """Send detailed order information to admin"""
    try:
        order_id = order_details['order_id']
        customer_name = order_details['customer_name'] or "Unknown"
        customer_phone = order_details['customer_phone']
        order_type = order_details['order_type'] or "GENERAL"
        total_amount = order_details['total_amount'] or 0
        status = order_details['status'] or "NEW"
        timestamp = order_details['timestamp']
        order_text = order_details['order_text'] or "N/A"
        
        status_emoji = {
            "NEW": "🆕",
            "PROCESSING": "🔄", 
            "COMPLETED": "✅",
            "CANCELLED": "❌"
        }.get(status, "📋")
        
        message = f"""📋 **Order Details**

**🆔 Order ID:** {order_id}
**👤 Customer:** {customer_name}
**📞 Phone:** {customer_phone}
**📦 Type:** {order_type}
**💰 Amount:** ${total_amount:.2f}
**📊 Status:** {status_emoji} {status}
**🕐 Created:** {timestamp}

**📝 Order Text:**
{order_text}

**🛍️ Products:**"""

        # Add products if available
        if order_details.get('products'):
            for i, product in enumerate(order_details['products'], 1):
                title = product.get('title', 'Unknown Product')
                quantity = product.get('quantity', 1)
                price = product.get('price', 0)
                item_total = product.get('item_total', 0)
                retailer_id = product.get('retailer_id', 'N/A')
                
                message += f"""
{i}. **{title}**
   • Qty: {quantity} | Price: ${price:.2f}
   • Total: ${item_total:.2f}
   • ID: {retailer_id}"""
        
        # Add admin notes if available
        if order_details.get('admin_notes'):
            message += f"\n\n**📝 Admin Notes:**\n{order_details['admin_notes']}"
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_update_status", title="🔄 Update Status"),
                ReplyButton(id="admin_contact_customer", title="📞 Contact"),
                ReplyButton(id="admin_order_dashboard", title="⬅️ Dashboard"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to send order details")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading order details: {str(e)}")


def send_admin_help(phone_number: str):
    """Send admin help message with available commands."""
    help_message = """🔧 **SpectraX Admin Panel**

**Button Interface:**
Use the admin dashboard buttons for easy management, or use text commands below.

**Text Commands:**

📋 **Management:**
• `/list` - View all current retailer IDs
• `/add_laptop <id>` - Add new laptop retailer ID
• `/add_repair <id>` - Add new repair retailer ID
• `/remove_laptop <id>` - Remove laptop retailer ID
• `/remove_repair <id>` - Remove repair retailer ID

📊 **Current Status:**
• You receive all order notifications
• Changes update Excel files automatically
• Changes take effect immediately

**Example Usage:**
• `/add_laptop abc123xyz` - Adds abc123xyz to laptop catalog
• `/add_repair def456uvw` - Adds def456uvw to repair catalog
• `/list` - Shows all current IDs

💡 **Tip:** Use the buttons below for easier navigation!"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=help_message,
        buttons=[
            ReplyButton(id="admin_manage_catalog", title="📝 Catalog"),
            ReplyButton(id="admin_view_stats", title="📊 View IDs"),
            ReplyButton(id="browse_laptops", title="👀 Preview"),
        ],
    )


def add_laptop_retailer_id(phone_number: str, retailer_id: str):
    """Add a new laptop retailer ID to the Excel file."""
    try:
        from catalog_utils import load_laptop_retailer_ids
        current_ids = load_laptop_retailer_ids()
        
        if retailer_id in current_ids:
            whatsapp.send_text(to=phone_number, body=f"⚠️ Laptop retailer ID '{retailer_id}' already exists!")
            return
        
        # Add to Excel file
        current_ids.append(retailer_id)
        update_laptop_excel(current_ids)
        
        whatsapp.send_text(to=phone_number, body=f"✅ Successfully added laptop retailer ID: {retailer_id}\n\nTotal laptop IDs: {len(current_ids)}")
        logger.info("Admin %s added laptop retailer ID: %s", phone_number, retailer_id)
        
    except Exception as e:
        logger.exception("Failed to add laptop retailer ID")
        whatsapp.send_text(to=phone_number, body=f"❌ Error adding laptop retailer ID: {str(e)}")


def add_repair_retailer_id(phone_number: str, retailer_id: str):
    """Add a new repair retailer ID to the Excel file."""
    try:
        from catalog_utils import load_repair_retailer_ids
        current_ids = load_repair_retailer_ids()
        
        if retailer_id in current_ids:
            whatsapp.send_text(to=phone_number, body=f"⚠️ Repair retailer ID '{retailer_id}' already exists!")
            return
        
        # Add to Excel file
        current_ids.append(retailer_id)
        update_repair_excel(current_ids)
        
        whatsapp.send_text(to=phone_number, body=f"✅ Successfully added repair retailer ID: {retailer_id}\n\nTotal repair IDs: {len(current_ids)}")
        logger.info("Admin %s added repair retailer ID: %s", phone_number, retailer_id)
        
    except Exception as e:
        logger.exception("Failed to add repair retailer ID")
        whatsapp.send_text(to=phone_number, body=f"❌ Error adding repair retailer ID: {str(e)}")


def remove_laptop_retailer_id(phone_number: str, retailer_id: str):
    """Remove a laptop retailer ID from the Excel file."""
    try:
        from catalog_utils import load_laptop_retailer_ids
        current_ids = load_laptop_retailer_ids()
        
        if retailer_id not in current_ids:
            whatsapp.send_text(to=phone_number, body=f"⚠️ Laptop retailer ID '{retailer_id}' not found!")
            return
        
        # Remove from list
        current_ids.remove(retailer_id)
        update_laptop_excel(current_ids)
        
        whatsapp.send_text(to=phone_number, body=f"✅ Successfully removed laptop retailer ID: {retailer_id}\n\nRemaining laptop IDs: {len(current_ids)}")
        logger.info("Admin %s removed laptop retailer ID: %s", phone_number, retailer_id)
        
    except Exception as e:
        logger.exception("Failed to remove laptop retailer ID")
        whatsapp.send_text(to=phone_number, body=f"❌ Error removing laptop retailer ID: {str(e)}")


def remove_repair_retailer_id(phone_number: str, retailer_id: str):
    """Remove a repair retailer ID from the Excel file."""
    try:
        from catalog_utils import load_repair_retailer_ids
        current_ids = load_repair_retailer_ids()
        
        if retailer_id not in current_ids:
            whatsapp.send_text(to=phone_number, body=f"⚠️ Repair retailer ID '{retailer_id}' not found!")
            return
        
        # Remove from list
        current_ids.remove(retailer_id)
        update_repair_excel(current_ids)
        
        whatsapp.send_text(to=phone_number, body=f"✅ Successfully removed repair retailer ID: {retailer_id}\n\nRemaining repair IDs: {len(current_ids)}")
        logger.info("Admin %s removed repair retailer ID: %s", phone_number, retailer_id)
        
    except Exception as e:
        logger.exception("Failed to remove repair retailer ID")
        whatsapp.send_text(to=phone_number, body=f"❌ Error removing repair retailer ID: {str(e)}")


def list_current_retailer_ids(phone_number: str):
    """List all current retailer IDs for admin."""
    try:
        from catalog_utils import load_laptop_retailer_ids, load_repair_retailer_ids
        laptop_ids = load_laptop_retailer_ids()
        repair_ids = load_repair_retailer_ids()
        
        laptop_list = "\n".join([f"  • {rid}" for rid in laptop_ids]) if laptop_ids else "  (none)"
        repair_list = "\n".join([f"  • {rid}" for rid in repair_ids]) if repair_ids else "  (none)"
        
        message = f"""📋 **Current Retailer IDs**

💻 **Laptops ({len(laptop_ids)} total):**
{laptop_list}

🛠 **Repairs ({len(repair_ids)} total):**
{repair_list}

Use `/add_laptop <id>` or `/add_repair <id>` to add more.
Use `/remove_laptop <id>` or `/remove_repair <id>` to remove."""
        
        whatsapp.send_text(to=phone_number, body=message)
        
    except Exception as e:
        logger.exception("Failed to list retailer IDs")
        whatsapp.send_text(to=phone_number, body=f"❌ Error listing retailer IDs: {str(e)}")


def update_laptop_excel(laptop_ids: List[str]):
    """Update the laptops.xlsx file with new laptop IDs."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add header
    ws.append(["retailer_id"])
    
    # Add laptop IDs
    for rid in laptop_ids:
        ws.append([rid])
    
    wb.save("laptops.xlsx")


def update_repair_excel(repair_ids: List[str]):
    """Update the repairs.xlsx file with new repair IDs."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add header
    ws.append(["retailer_id"])
    
    # Add repair IDs
    for rid in repair_ids:
        ws.append([rid])
    
    wb.save("repairs.xlsx")


@app.get("/")
def read_root():
    return {"message": "Welcome to SpectraX Laptops WhatsApp Bot!"}


@app.get("/webhook")
async def verify_webhook(request: Request):
    params = request.query_params
    mode = params.get("hub.mode")
    token = params.get("hub.verify_token")
    challenge = params.get("hub.challenge")

    logger.info("Verifying webhook: mode=%s, token=%s", mode, token)

    if token and token == VERIFY_TOKEN:
        logger.info("Token verified successfully.")
        return PlainTextResponse(content=challenge)

    logger.warning("Invalid verify token: %s", token)
    raise HTTPException(status_code=403, detail="Invalid verify token")


@app.post("/webhook")
async def receive_message(request: Request):
    try:
        body = await request.body()
        message = whatsapp.parse(body)
        
        # Generate session ID for this interaction
        session_id = generate_session_id()

        if isinstance(message, TextMessage):
            # Mark the incoming message as read (safe)
            safe_mark_as_read(message.id)
            
            # Extract text content safely and check admin commands first
            _text = _get_text_content(message)
            user_name = getattr(message.user, 'name', 'Unknown')
            phone_number = message.user.phone_number
            is_admin_user = is_admin(phone_number)
            
            # Log the incoming text message
            activity_logger.log_activity(
                phone_number=phone_number,
                user_name=user_name,
                activity_type="message_received",
                message_type="text",
                user_input=_text,
                admin_flag=is_admin_user,
                session_id=session_id
            )
            
            if _text and handle_admin_command(phone_number, _text):
                # Log admin command execution
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="admin_command",
                    message_type="text",
                    user_input=_text,
                    bot_response="Admin command processed",
                    admin_flag=True,
                    session_id=session_id
                )
                return {"status": "admin_command_processed"}
            
            # Check if it's admin - send admin welcome instead of regular welcome
            if is_admin_user:
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="admin_welcome",
                    message_type="text",
                    bot_response="Admin welcome message sent",
                    admin_flag=True,
                    session_id=session_id
                )
                send_admin_welcome_message(phone_number)
            else:
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="welcome_message",
                    message_type="text",
                    bot_response="Welcome message sent",
                    admin_flag=False,
                    session_id=session_id
                )
                send_welcome_message(phone_number)

        elif isinstance(message, InteractiveButtonMessage):
            # Mark the incoming message as read (safe)
            safe_mark_as_read(message.id)
            user_choice = message.reply_id
            phone_number = message.user.phone_number
            user_name = getattr(message.user, 'name', 'Unknown')
            is_admin_user = is_admin(phone_number)
            
            # Log the button click
            activity_logger.log_activity(
                phone_number=phone_number,
                user_name=user_name,
                activity_type="button_clicked",
                message_type="interactive_button",
                user_input=f"Button: {user_choice}",
                button_id=user_choice,
                admin_flag=is_admin_user,
                session_id=session_id
            )
            
            if user_choice == "browse_laptops":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="browse_laptops",
                    bot_response="Laptop catalog options sent",
                    button_id=user_choice,
                    admin_flag=is_admin_user,
                    session_id=session_id
                )
                handle_browse_laptops(phone_number)
            elif user_choice == "browse_collection":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="browse_collection",
                    bot_response="Collection browsing initiated",
                    button_id=user_choice,
                    admin_flag=is_admin_user,
                    session_id=session_id
                )
                handle_browse_laptops(phone_number)
            elif user_choice == "why_spectrax":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="why_spectrax",
                    bot_response="Why SpectraX message sent",
                    button_id=user_choice,
                    admin_flag=is_admin_user,
                    session_id=session_id
                )
                send_why_spectrax_message(phone_number)
            elif user_choice == "lifetime_support":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="lifetime_support",
                    bot_response="Lifetime support info sent",
                    button_id=user_choice,
                    admin_flag=is_admin_user,
                    session_id=session_id
                )
                send_lifetime_support_message(phone_number)
            elif user_choice == "see_collection_from_why":
                handle_browse_laptops(phone_number)
            elif user_choice == "support_from_why":
                send_lifetime_support_message(phone_number)
            elif user_choice == "browse_from_support":
                handle_browse_laptops(phone_number)
            elif user_choice == "how_to_order":
                send_how_to_order_message(phone_number)
            elif user_choice == "register_laptop":
                send_registration_flow(phone_number)
            elif user_choice == "schedule_service":
                send_service_booking_flow(phone_number)
            elif user_choice == "request_video_demo":
                # Run video demo as background task to avoid blocking the webhook response
                asyncio.create_task(handle_video_demo_request(phone_number))
            elif user_choice == "upgrades_accessories":
                send_upgrades_accessories_message(phone_number)
            # new button handlers
            elif user_choice == "action_buy_laptop":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="catalog_viewed",
                    bot_response="Laptop catalog sent",
                    button_id=user_choice,
                    admin_flag=is_admin_user,
                    session_id=session_id,
                    additional_data={"catalog_type": "laptops"}
                )
                handle_buy_laptops(phone_number)
            elif user_choice == "action_repairs":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="catalog_viewed",
                    bot_response="Repair catalog sent",
                    button_id=user_choice,
                    admin_flag=is_admin_user,
                    session_id=session_id,
                    additional_data={"catalog_type": "repairs"}
                )
                handle_repairs(phone_number)
            # Admin button handlers
            elif user_choice == "admin_catalog_management":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="admin_catalog_management",
                    bot_response="Catalog management menu sent",
                    button_id=user_choice,
                    admin_flag=True,
                    session_id=session_id
                )
                send_admin_catalog_menu(phone_number)
            elif user_choice == "admin_order_management":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="admin_order_management",
                    bot_response="Order management menu sent",
                    button_id=user_choice,
                    admin_flag=True,
                    session_id=session_id
                )
                send_admin_order_menu(phone_number)
            elif user_choice == "admin_manage_catalog":
                send_admin_catalog_menu(phone_number)
            elif user_choice == "admin_view_stats":
                list_current_retailer_ids(phone_number)
            elif user_choice == "admin_add_laptop":
                send_add_laptop_prompt(phone_number)
            elif user_choice == "admin_add_repair":
                send_add_repair_prompt(phone_number)
            elif user_choice == "admin_remove_laptop":
                send_remove_laptop_menu(phone_number)
            elif user_choice == "admin_remove_repair":
                send_remove_repair_menu(phone_number)
            elif user_choice == "admin_back_main":
                if is_admin(phone_number):
                    send_admin_welcome_message(phone_number)
                else:
                    send_welcome_message(phone_number)
            # Order management handlers (placeholders for now)
            elif user_choice == "admin_recent_orders":
                send_admin_recent_orders(phone_number)
            elif user_choice == "admin_order_status":
                send_admin_order_status_menu(phone_number)
            elif user_choice == "admin_customer_comm":
                send_admin_customer_comm_menu(phone_number)
            elif user_choice == "admin_order_analytics":
                send_admin_order_analytics(phone_number)
            elif user_choice == "admin_delivery_tracking":
                send_admin_delivery_tracking(phone_number)
            elif user_choice == "admin_activity_stats":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="admin_activity_stats",
                    bot_response="Activity statistics sent",
                    button_id=user_choice,
                    admin_flag=True,
                    session_id=session_id
                )
                send_admin_activity_stats(phone_number)
            elif user_choice == "admin_analytics_menu":
                activity_logger.log_activity(
                    phone_number=phone_number,
                    user_name=user_name,
                    activity_type="admin_analytics_menu",
                    bot_response="Analytics menu sent",
                    button_id=user_choice,
                    admin_flag=True,
                    session_id=session_id
                )
                send_admin_analytics_menu(phone_number)
            elif user_choice == "admin_detailed_analytics":
                send_admin_detailed_analytics(phone_number)
            elif user_choice == "admin_conversation_analytics":
                send_admin_conversation_analytics(phone_number)
            elif user_choice == "admin_export_menu":
                send_admin_export_menu(phone_number)
            elif user_choice == "admin_export_data":
                send_admin_export_menu(phone_number)
            elif user_choice == "admin_export_7days":
                handle_admin_export_request(phone_number, "7days")
            elif user_choice == "admin_export_30days":
                handle_admin_export_request(phone_number, "30days")
            elif user_choice == "admin_export_admin_only":
                handle_admin_export_request(phone_number, "admin_only")
            elif user_choice == "admin_export_conversations":
                handle_admin_export_request(phone_number, "conversations")
            
            # Order processing handlers
            elif user_choice == "admin_process_order":
                send_admin_order_processing_menu(phone_number)
            elif user_choice == "admin_contact_customer":
                send_admin_contact_customer_menu(phone_number)
            elif user_choice == "admin_order_details":
                send_admin_order_details_menu(phone_number)
            elif user_choice == "admin_mark_processing":
                whatsapp.send_text(to=phone_number, body="✅ Order marked as processing. Customer will be notified of status update.")
            elif user_choice == "admin_request_payment":
                whatsapp.send_text(to=phone_number, body="💳 Payment request template sent to customer. Follow up via phone for confirmation.")
            elif user_choice == "admin_schedule_delivery":
                whatsapp.send_text(to=phone_number, body="🚚 Delivery scheduling initiated. Contact customer to confirm preferred time slots.")
            elif user_choice == "admin_send_confirmation":
                whatsapp.send_text(to=phone_number, body="✅ Order confirmation sent to customer with details and next steps.")
            elif user_choice == "admin_request_details":
                whatsapp.send_text(to=phone_number, body="📝 Additional details request sent to customer. Await response for order processing.")
            elif user_choice == "admin_schedule_call":
                whatsapp.send_text(to=phone_number, body="📞 Call scheduled with customer. Follow up within agreed timeframe.")
            elif user_choice == "admin_view_all_orders":
                whatsapp.send_text(to=phone_number, body="📊 Displaying recent orders. Check activity log for complete order history.")
            elif user_choice == "admin_update_status":
                whatsapp.send_text(to=phone_number, body="🔄 Order status update interface. Select order to modify status.")

        elif isinstance(message, OrderMessage):
            # Mark order message as read (safe)
            safe_mark_as_read(message.id)
            
            user_name = getattr(message.user, 'name', 'Unknown')
            phone_number = message.user.phone_number

            # Prepare order details for logging
            order_details = {
                "catalog_id": message.catalog_id,
                "order_text": message.order_text,
                "products": []
            }

            try:
                from catalog_utils import load_laptop_retailer_ids, load_repair_retailer_ids
                laptop_ids = load_laptop_retailer_ids()
                repair_ids = load_repair_retailer_ids()
                
                # Check product retailer IDs to determine order type
                order_retailer_ids = []
                total_amount = 0
                
                for product in message.products:
                    # Try multiple field names for better product data extraction
                    product_title = (
                        getattr(product, "title", None)
                        or getattr(product, "name", None) 
                        or getattr(product, "product_name", None)
                        or getattr(product, "description", None)
                        or "Unnamed Product"
                    )
                    
                    product_id = (
                        getattr(product, "product_retailer_id", None)
                        or getattr(product, "product_id", None)
                        or getattr(product, "id", None)
                        or "N/A"
                    )
                    
                    quantity = getattr(product, "quantity", getattr(product, "quantity_ordered", 1))
                    price = getattr(product, "retail_price", getattr(product, "price", 0))
                    
                    # Calculate item total
                    try:
                        price_float = float(str(price).replace('$', '').replace(',', ''))
                        item_total = price_float * quantity
                        total_amount += item_total
                    except:
                        price_float = 0
                        item_total = 0
                    
                    product_data = {
                        "title": product_title,
                        "quantity": quantity,
                        "price": price_float,
                        "item_total": item_total,
                        "retailer_id": product_id
                    }
                    order_details["products"].append(product_data)
                    
                    if hasattr(product, 'product_retailer_id') and product.product_retailer_id:
                        order_retailer_ids.append(product.product_retailer_id)
                    elif product_id != "N/A":
                        order_retailer_ids.append(product_id)
                
                # Determine if it's laptops, repairs, or mixed
                laptop_count = sum(1 for rid in order_retailer_ids if rid in laptop_ids)
                repair_count = sum(1 for rid in order_retailer_ids if rid in repair_ids)
                
                if laptop_count > 0 and repair_count == 0:
                    order_type = "LAPTOP"
                elif repair_count > 0 and laptop_count == 0:
                    order_type = "REPAIR"
                elif laptop_count > 0 and repair_count > 0:
                    order_type = "MIXED"
                else:
                    order_type = "GENERAL"
                    
                order_details["order_type"] = order_type
                order_details["laptop_count"] = laptop_count
                order_details["repair_count"] = repair_count
                order_details["total_amount"] = total_amount
                
            except Exception as e:
                logger.exception("Failed to determine order type: %s", e)
                order_type = "UNKNOWN"
                total_amount = 0

            # Log the order to Excel
            try:
                order_id = order_logger.log_order(
                    customer_phone=phone_number,
                    customer_name=user_name,
                    order_type=order_type,
                    total_amount=total_amount,
                    catalog_id=message.catalog_id,
                    order_text=message.order_text,
                    products_data=order_details["products"],
                    currency="USD",
                    status="NEW"
                )
                logger.info(f"Order logged with ID: {order_id}")
            except Exception as e:
                logger.exception("Failed to log order to Excel: %s", e)
                order_id = f"ERR_{datetime.now().strftime('%Y%m%d%H%M%S')}"

            # Log the order activity
            activity_logger.log_activity(
                phone_number=phone_number,
                user_name=user_name,
                activity_type="order_placed",
                message_type="order",
                user_input=f"Order placed: {order_type}",
                bot_response="Order confirmation sent",
                admin_flag=False,
                session_id=session_id,
                additional_data={**order_details, "order_id": order_id}
            )

            # Build enhanced order summary for admin notification
            admin_summary_lines = [
                f"🚨 **NEW {order_type} ORDER RECEIVED**",
                f"🆔 Order ID: {order_id}",
                f"📱 Customer: {user_name} ({phone_number})",
                f"🕐 Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"📋 Order Text: {message.order_text}",
                f"🏪 Catalog ID: {message.catalog_id}",
                f"� Total Amount: ${total_amount:.2f}",
                "",
                "📦 **PRODUCTS ORDERED:**",
            ]
            
            for i, product_data in enumerate(order_details["products"], 1):
                admin_summary_lines.extend([
                    f"{i}. **{product_data['title']}**",
                    f"   • Quantity: {product_data['quantity']}",
                    f"   • Unit Price: ${product_data['price']:.2f}",
                    f"   • Subtotal: ${product_data['item_total']:.2f}",
                    f"   • Product ID: {product_data['retailer_id']}",
                    ""
                ])
            
            # Add order totals and summary
            admin_summary_lines.extend([
                "💰 **ORDER SUMMARY:**",
                f"• Order ID: {order_id}",
                f"• Total Items: {len(message.products)}",
                f"• Total Amount: ${total_amount:.2f}",
                f"• Order Type: {order_type}",
                f"• Status: NEW",
                "",
                "⚡ **NEXT ACTIONS:**",
                "1. Contact customer within 30 minutes",
                "2. Confirm payment method & delivery details",
                "3. Process inventory and prepare shipment",
                "4. Update order status in system",
                "",
                f"� **Customer Contact:** {phone_number}",
                f"� **Customer Name:** {user_name}",
            ])
            
            admin_message = "\n".join(admin_summary_lines)
            
            # Send detailed order notification to admin
            try:
                whatsapp.send_interactive_buttons(
                    to=ADMIN_NUMBER,
                    body=admin_message,
                    buttons=[
                        ReplyButton(id="admin_process_order", title="✅ Process Order"),
                        ReplyButton(id="admin_contact_customer", title="📞 Contact"),
                        ReplyButton(id="admin_order_details", title="📋 Details"),
                    ],
                )
                
                # Also log the admin notification
                activity_logger.log_activity(
                    phone_number=ADMIN_NUMBER,
                    user_name="System",
                    activity_type="admin_order_notification",
                    message_type="system",
                    bot_response=f"Order notification sent for {order_type} order from {phone_number}",
                    admin_flag=True,
                    session_id=session_id,
                    additional_data={
                        "customer_phone": phone_number,
                        "customer_name": user_name,
                        "order_type": order_type,
                        "order_value": total_amount,
                        "product_count": len(message.products)
                    }
                )
                
            except Exception as admin_error:
                logger.exception("Failed to send order notification to admin: %s", admin_error)

            # Send confirmation to customer based on order type
            if order_type in ["REPAIR", "MIXED (LAPTOP + REPAIR)"]:
                customer_response = f"""🎉 Awesome! We've received your {order_type.lower()} order!

**What happens next:**
1️⃣ Our team will contact you within 30 minutes
2️⃣ Confirm service details & payment method
3️⃣ Schedule pickup/drop-off for repairs
4️⃣ Complete service registration to unlock:
   • FREE diagnosis & quote
   • Lifetime repair tracking  
   • WhatsApp progress updates

🛠 **Remember**: Service registration enables tracking and priority support!

Thanks for choosing SpectraX Laptop Services! 🔧✨"""
            else:
                customer_response = f"""🎉 Awesome! We've received your {order_type.lower()} order!

**What happens next:**
1️⃣ Our team will contact you within 30 minutes
2️⃣ Confirm payment method & delivery details  
3️⃣ Schedule delivery & setup if needed
4️⃣ Complete laptop registration to unlock:
   • FREE Starter Essentials software
   • Lifetime repair tracking
   • WhatsApp tech support

🎁 **Remember**: Registration unlocks amazing benefits, so don't skip this step!

Thanks for choosing SpectraX Laptops! 💻✨"""
            
            # Send customer confirmation
            whatsapp.send_text(
                to=message.user.phone_number,
                body=customer_response
            )

            # Log customer confirmation
            activity_logger.log_activity(
                phone_number=phone_number,
                user_name=user_name,
                activity_type="order_confirmation_sent",
                message_type="text",
                bot_response=f"Order confirmation sent for {order_type}",
                admin_flag=False,
                session_id=session_id
            )

        return {"status": "processed"}
    except Exception as e:
        logger.error("Error processing message: %s", str(e))
        return {"status": "error", "message": str(e)}


def send_welcome_message(phone_number):
    """Send the initial welcome message with quick reply buttons for laptop offerings"""
    message = """� Welcome to SpectraX Laptops!  
Your trusted partner for premium laptops with lifetime support 🚀  

🎁 **Special Launch Offer**: Buy any laptop → get FREE Starter Essentials software + lifetime repair tracking when registered!

Choose an option below 👇"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="browse_laptops", title="💻 Laptops"),
            ReplyButton(id="why_spectrax", title="💡 Why Us?"),
            ReplyButton(id="lifetime_support", title="🛡 Support"),
        ],
    )


def send_admin_order_dashboard(phone_number: str):
    """Send order management dashboard with real data from Excel"""
    try:
        # Get order statistics
        stats = order_logger.get_order_statistics()
        recent_orders = order_logger.get_recent_orders(5)
        
        message = f"""📋 **Order Management Dashboard**

**📊 Order Statistics:**
• Total Orders: {stats['total_orders']}
• New Orders: {stats['new_orders']}
• Processing: {stats['processing_orders']}
• Completed: {stats['completed_orders']}
• Cancelled: {stats['cancelled_orders']}

**💰 Revenue:**
• Total Revenue: ${stats['total_revenue']:.2f}
• Avg Order Value: ${stats['average_order_value']:.2f}

**🕐 Recent Orders:**"""

        if recent_orders:
            for order in recent_orders[:3]:
                order_id = order['order_id'][:12] + "..." if len(order['order_id']) > 12 else order['order_id']
                customer = order['customer_name'] or "Unknown"
                amount = f"${order['total_amount']:.2f}" if order['total_amount'] else "N/A"
                status = order['status'] or "NEW"
                
                status_emoji = {
                    "NEW": "🆕",
                    "PROCESSING": "🔄", 
                    "COMPLETED": "✅",
                    "CANCELLED": "❌"
                }.get(status, "📋")
                
                message += f"\n{status_emoji} {order_id} - {customer} - {amount}"
        else:
            message += "\nNo recent orders found."
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_new_orders", title="🆕 New Orders"),
                ReplyButton(id="admin_all_orders", title="📋 All Orders"),
                ReplyButton(id="admin_order_search", title="🔍 Search Order"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to get order dashboard")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading order dashboard: {str(e)}")


def send_admin_new_orders(phone_number: str):
    """Show all new orders that need processing"""
    try:
        new_orders = order_logger.get_orders_by_status("NEW")
        
        if not new_orders:
            message = "🆕 **New Orders**\n\nNo new orders found! All orders have been processed. 🎉"
        else:
            message = f"🆕 **New Orders ({len(new_orders)})**\n\n"
            
            for order in new_orders[-10:]:  # Show last 10 new orders
                order_id = order['order_id']
                customer = order['customer_name'] or "Unknown"
                customer_phone = order['customer_phone'][-4:] if order['customer_phone'] else "N/A"
                amount = f"${order['total_amount']:.2f}" if order['total_amount'] else "N/A"
                order_type = order['order_type'] or "GENERAL"
                timestamp = order['timestamp']
                
                message += f"""**{order_id}**
👤 {customer} (...{customer_phone})
💰 {amount} | 📦 {order_type}
🕐 {timestamp}

"""
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_process_next", title="⚡ Process Next"),
                ReplyButton(id="admin_order_details", title="📋 Order Details"),
                ReplyButton(id="admin_order_dashboard", title="⬅️ Dashboard"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to get new orders")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading new orders: {str(e)}")


def send_admin_all_orders(phone_number: str):
    """Show all orders with filtering options"""
    try:
        all_orders = order_logger.get_recent_orders(20)  # Last 20 orders
        stats = order_logger.get_order_statistics()
        
        message = f"""📋 **All Orders (Last 20)**

**Quick Stats:**
🆕 New: {stats['new_orders']} | 🔄 Processing: {stats['processing_orders']}
✅ Completed: {stats['completed_orders']} | ❌ Cancelled: {stats['cancelled_orders']}

**Recent Orders:**
"""
        
        if all_orders:
            for order in all_orders:
                order_id = order['order_id'][:10] + "..." if len(order['order_id']) > 10 else order['order_id']
                customer = (order['customer_name'] or "Unknown")[:15]
                amount = f"${order['total_amount']:.2f}" if order['total_amount'] else "N/A"
                status = order['status'] or "NEW"
                
                status_emoji = {
                    "NEW": "🆕",
                    "PROCESSING": "🔄", 
                    "COMPLETED": "✅",
                    "CANCELLED": "❌"
                }.get(status, "📋")
                
                message += f"{status_emoji} {order_id} | {customer} | {amount}\n"
        else:
            message += "No orders found."
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_filter_status", title="🔍 Filter Status"),
                ReplyButton(id="admin_export_orders", title="📥 Export Orders"),
                ReplyButton(id="admin_order_dashboard", title="⬅️ Dashboard"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to get all orders")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading all orders: {str(e)}")


def handle_admin_order_update(phone_number: str, message_text: str):
    """Handle order status updates via text commands"""
    try:
        text = message_text.lower().strip()
        
        # Format: /update_order ORDER_ID STATUS [NOTES]
        if text.startswith("/update_order "):
            parts = message_text[14:].strip().split(' ', 2)
            if len(parts) < 2:
                whatsapp.send_text(
                    to=phone_number, 
                    body="❌ Invalid format. Use: /update_order ORDER_ID STATUS [NOTES]\n\nValid statuses: NEW, PROCESSING, COMPLETED, CANCELLED"
                )
                return True
            
            order_id = parts[0].upper()
            status = parts[1].upper()
            notes = parts[2] if len(parts) > 2 else ""
            
            # Validate status
            valid_statuses = ["NEW", "PROCESSING", "COMPLETED", "CANCELLED"]
            if status not in valid_statuses:
                whatsapp.send_text(
                    to=phone_number,
                    body=f"❌ Invalid status. Valid statuses: {', '.join(valid_statuses)}"
                )
                return True
            
            # Update order
            success = order_logger.update_order_status(order_id, status, notes, "Admin")
            
            if success:
                # Get order details for confirmation
                order_details = order_logger.get_order_details(order_id)
                if order_details:
                    customer_name = order_details['customer_name'] or "Unknown"
                    customer_phone = order_details['customer_phone']
                    
                    confirmation_message = f"""✅ **Order Updated Successfully**

**Order ID:** {order_id}
**New Status:** {status}
**Customer:** {customer_name}
**Notes Added:** {notes or "None"}

**Next Steps:**
• Notify customer of status change
• Update inventory if completed
• Process refund if cancelled"""
                    
                    whatsapp.send_interactive_buttons(
                        to=phone_number,
                        body=confirmation_message,
                        buttons=[
                            ReplyButton(id="admin_notify_customer", title="📞 Notify Customer"),
                            ReplyButton(id="admin_order_details", title="📋 Order Details"),
                            ReplyButton(id="admin_new_orders", title="🆕 New Orders"),
                        ],
                    )
                    
                    # Log the admin action
                    activity_logger.log_activity(
                        phone_number=phone_number,
                        user_name="Admin",
                        activity_type="admin_order_update",
                        message_type="text",
                        user_input=message_text,
                        bot_response=f"Order {order_id} updated to {status}",
                        admin_flag=True,
                        additional_data={
                            "order_id": order_id,
                            "old_status": order_details.get('status', 'Unknown'),
                            "new_status": status,
                            "notes": notes
                        }
                    )
                else:
                    whatsapp.send_text(to=phone_number, body="✅ Order updated but couldn't retrieve details.")
            else:
                whatsapp.send_text(to=phone_number, body=f"❌ Failed to update order {order_id}. Order may not exist.")
            
            return True
        
        return False
        
    except Exception as e:
        logger.exception("Failed to handle order update")
        whatsapp.send_text(to=phone_number, body=f"❌ Error updating order: {str(e)}")
        return True


def send_admin_welcome_message(phone_number: str):
    """Send admin welcome message with management options"""
    message = """🔧 **SpectraX Admin Dashboard**

Welcome back, Admin! 👋

**Quick Stats:**
"""
    
    try:
        from catalog_utils import load_laptop_retailer_ids, load_repair_retailer_ids
        laptop_count = len(load_laptop_retailer_ids())
        repair_count = len(load_repair_retailer_ids())
        
        # Get activity stats
        recent_activities = activity_logger.get_recent_activities(5)
        today_activities = len([a for a in recent_activities if str(a['timestamp']).startswith(datetime.now().strftime("%Y-%m-%d"))])
        total_conversations = len(set(activity['phone_number'] for activity in recent_activities))
        
        message += f"💻 Laptop Products: {laptop_count}\n"
        message += f"🛠 Repair Services: {repair_count}\n"
        message += f"� Today's Activities: {today_activities}\n"
        message += f"�💬 Recent Users: {total_conversations}\n"
        message += f"⏰ Last Activity: {recent_activities[0]['timestamp'] if recent_activities else 'None'}\n\n"
    except:
        message += "📊 Loading statistics...\n\n"
    
    message += "**Management Areas:**"
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_catalog_management", title="📝 Catalog"),
            ReplyButton(id="admin_order_management", title="📦 Orders"),
            ReplyButton(id="admin_analytics_menu", title="📊 Analytics"),
        ],
    )


def send_admin_catalog_menu(phone_number: str):
    """Send admin catalog management menu"""
    message = """📝 **Catalog Management**

Manage your product catalog:

**Product Management:**
• Add new laptop retailer IDs
• Add new repair service IDs
• Remove existing products
• View all current products

**Quick Actions:**"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_add_laptop", title="➕ Laptop"),
            ReplyButton(id="admin_add_repair", title="➕ Repair"),
            ReplyButton(id="admin_remove_laptop", title="➖ Laptop"),
        ],
    )
    
    # Send second set of buttons
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body="**More Options:**",
        buttons=[
            ReplyButton(id="admin_remove_repair", title="➖ Repair"),
            ReplyButton(id="admin_view_stats", title="📊 View IDs"),
            ReplyButton(id="admin_back_main", title="⬅️ Main"),
        ],
    )


def send_admin_order_menu(phone_number: str):
    """Send admin order management menu"""
    message = """📦 **Order Management**

Manage customer orders and services:

**Order Status:**
• View recent orders
• Update order status
• Track deliveries
• Manage repairs

**Customer Communication:**
• Send status updates
• Handle inquiries
• Process refunds"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_recent_orders", title="📋 Orders"),
            ReplyButton(id="admin_order_status", title="🔄 Status"),
            ReplyButton(id="admin_customer_comm", title="💬 Customer"),
        ],
    )
    
    # Send second set of buttons
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body="**More Options:**",
        buttons=[
            ReplyButton(id="admin_order_analytics", title="📊 Analytics"),
            ReplyButton(id="admin_delivery_tracking", title="🚚 Delivery"),
            ReplyButton(id="admin_back_main", title="⬅️ Main"),
        ],
    )


def send_add_laptop_prompt(phone_number: str):
    """Prompt admin to add laptop retailer ID"""
    message = """➕ **Add Laptop Retailer ID**

To add a new laptop to the catalog, reply with:
`/add_laptop <retailer_id>`

**Example:**
`/add_laptop new_laptop_123`

The new laptop will be immediately available in the catalog! 🚀"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_manage_catalog", title="⬅️ Back to Menu"),
            ReplyButton(id="admin_view_stats", title="📊 View Current IDs"),
        ],
    )


def send_add_repair_prompt(phone_number: str):
    """Prompt admin to add repair retailer ID"""
    message = """➕ **Add Repair Service ID**

To add a new repair service to the catalog, reply with:
`/add_repair <retailer_id>`

**Example:**
`/add_repair new_repair_456`

The new repair service will be immediately available! 🛠"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_manage_catalog", title="⬅️ Back to Menu"),
            ReplyButton(id="admin_view_stats", title="📊 View Current IDs"),
        ],
    )


def send_remove_laptop_menu(phone_number: str):
    """Send menu to remove laptop retailer IDs"""
    try:
        from catalog_utils import load_laptop_retailer_ids
        laptop_ids = load_laptop_retailer_ids()
        
        if not laptop_ids:
            message = "ℹ️ **No Laptop IDs to Remove**\n\nThere are currently no laptop retailer IDs in the system."
            whatsapp.send_interactive_buttons(
                to=phone_number,
                body=message,
                buttons=[ReplyButton(id="admin_manage_catalog", title="⬅️ Back to Menu")],
            )
            return
        
        laptop_list = "\n".join([f"• {rid}" for rid in laptop_ids])
        message = f"""➖ **Remove Laptop Retailer ID**

**Current Laptop IDs:**
{laptop_list}

To remove a laptop, reply with:
`/remove_laptop <retailer_id>`

**Example:**
`/remove_laptop {laptop_ids[0]}`"""
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_manage_catalog", title="⬅️ Back to Menu"),
                ReplyButton(id="admin_view_stats", title="📊 View All IDs"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to load laptop IDs for removal")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading laptop IDs: {str(e)}")


def send_remove_repair_menu(phone_number: str):
    """Send menu to remove repair retailer IDs"""
    try:
        from catalog_utils import load_repair_retailer_ids
        repair_ids = load_repair_retailer_ids()
        
        if not repair_ids:
            message = "ℹ️ **No Repair IDs to Remove**\n\nThere are currently no repair retailer IDs in the system."
            whatsapp.send_interactive_buttons(
                to=phone_number,
                body=message,
                buttons=[ReplyButton(id="admin_catalog_management", title="⬅️ Back to Catalog")],
            )
            return
        
        repair_list = "\n".join([f"• {rid}" for rid in repair_ids])
        message = f"""➖ **Remove Repair Service ID**

**Current Repair IDs:**
{repair_list}

To remove a repair service, reply with:
`/remove_repair <retailer_id>`

**Example:**
`/remove_repair {repair_ids[0]}`"""
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_catalog_management", title="⬅️ Back to Catalog"),
                ReplyButton(id="admin_view_stats", title="📊 View All IDs"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to load repair IDs for removal")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading repair IDs: {str(e)}")


def send_admin_recent_orders(phone_number: str):
    """Send recent orders overview (placeholder)"""
    message = """📋 **Recent Orders**

**Last 24 Hours:**
• 3 Laptop Orders
• 2 Repair Services
• 1 Mixed Order

**Status Overview:**
✅ 4 Completed
🔄 2 Processing
📦 0 Pending

*Note: Full order management system coming soon!*"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_order_status", title="🔄 Update Status"),
            ReplyButton(id="admin_order_management", title="⬅️ Back to Orders"),
            ReplyButton(id="admin_back_main", title="🏠 Main Menu"),
        ],
    )


def send_admin_order_status_menu(phone_number: str):
    """Send order status update menu (placeholder)"""
    message = """🔄 **Update Order Status**

**Available Status Updates:**
• Order Received → Processing
• Processing → Shipped
• Shipped → Delivered
• Mark as Completed

**Instructions:**
Send order ID with new status to update.

*Example: ORDER123 shipped*

*Note: Advanced status tracking system in development!*"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_recent_orders", title="📋 View Orders"),
            ReplyButton(id="admin_order_management", title="⬅️ Back to Orders"),
        ],
    )


def send_admin_customer_comm_menu(phone_number: str):
    """Send customer communication menu (placeholder)"""
    message = """💬 **Customer Communication**

**Available Actions:**
• Send delivery updates
• Answer product inquiries
• Handle support requests
• Process feedback

**Quick Templates:**
• Order confirmation
• Shipping notification
• Delivery confirmation
• Service completion

*Note: Template system and automation coming soon!*"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_order_analytics", title="📊 View Analytics"),
            ReplyButton(id="admin_order_management", title="⬅️ Back to Orders"),
        ],
    )


def send_admin_order_analytics(phone_number: str):
    """Send order analytics overview (placeholder)"""
    message = """📊 **Order Analytics**

**This Week:**
📈 Total Orders: 15 (+25%)
💰 Revenue: $4,500 (+30%)
⭐ Avg Rating: 4.8/5

**Top Products:**
1. Gaming Laptops (40%)
2. Business Laptops (35%)
3. Repair Services (25%)

**Customer Satisfaction:**
😊 95% Positive Feedback
🔄 5% Return Rate

*Note: Advanced analytics dashboard in development!*"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_delivery_tracking", title="🚚 Delivery Status"),
            ReplyButton(id="admin_order_management", title="⬅️ Back to Orders"),
        ],
    )


def send_admin_delivery_tracking(phone_number: str):
    """Send delivery tracking overview (placeholder)"""
    message = """🚚 **Delivery Tracking**

**Active Deliveries:**
📦 ORDER123 - En route (ETA: 2 hours)
📦 ORDER124 - Preparing for dispatch
📦 ORDER125 - Out for delivery

**Delivery Stats:**
✅ 95% On-time delivery
🕐 Avg delivery time: 24 hours
📍 Coverage: All major cities

**Next Actions:**
• Update delivery status
• Contact delivery partner
• Handle delivery issues

*Note: Real-time tracking integration coming soon!*"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_customer_comm", title="💬 Customer Updates"),
            ReplyButton(id="admin_order_management", title="⬅️ Back to Orders"),
        ],
    )


def send_buy_repairs_buttons(phone_number: str):
    """Send two reply buttons: Buy Laptop and Repairs (reuses existing ReplyButton pattern)."""
    body = "Choose an option:\n\n🛒 Buy Laptop — view laptops to purchase\n🛠 Repairs — view repair offering"
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=body,
        buttons=[
            ReplyButton(id="action_buy_laptop", title="🛒 Buy Laptop"),
            ReplyButton(id="action_repairs", title="🛠 Repairs"),
        ],
    )


def handle_browse_laptops(phone_number):
    """Show Buy / Repairs choices instead of immediately sending the catalog."""
    send_buy_repairs_buttons(phone_number)


# Delegate catalog handling to separate modules
try:
    from laptops import handle_buy_laptops as _handle_buy_laptops_module
    from repairs import handle_repairs as _handle_repairs_module
except Exception:
    _handle_buy_laptops_module = None
    _handle_repairs_module = None


def handle_buy_laptops(phone_number: str):
    """Delegate to laptops module if available, otherwise fall back to inline implementation."""
    if _handle_buy_laptops_module:
        return _handle_buy_laptops_module(whatsapp, phone_number, catalog_id=CATALOG_ID)

    # Fallback: load laptop IDs from separate Excel file
    from catalog_utils import load_laptop_retailer_ids
    laptop_ids = load_laptop_retailer_ids()
    if not laptop_ids:
        laptop_ids = _env_retailer_ids("PRODUCT_RETAILER_ID", "PRODUCT_RETAILER_ID_2")

    if not laptop_ids:
        logger.warning("No laptop retailer IDs configured (env or excel)")
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body="No laptops are configured right now. Contact support to add products.",
            buttons=[ReplyButton(id="contact_support", title="Contact Support")],
        )
        return

    header = "SpectraX Laptop Catalog"
    body = "💻 Browse our featured laptops. Each purchase includes FREE Starter Essentials software + lifetime repair tracking when registered."
    footer = "Tap a laptop to view details & order."

    try:
        # Use the safe catalog compatibility function
        from catalog_utils import send_catalog_compat
        send_catalog_compat(
            whatsapp=whatsapp,
            to=phone_number,
            retailer_ids=laptop_ids,
            header=header,
            body=body,
            footer=footer,
            catalog_id=CATALOG_ID,
            fallback_button_id="browse_laptops"
        )
    except Exception as exc:
        logger.exception("Failed sending laptop catalog: %s", exc)
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body="Sorry, something went wrong while fetching the catalog. Try again later.",
            buttons=[ReplyButton(id="try_again", title="Try Again")],
        )


def handle_repairs(phone_number: str):
    """Delegate to repairs module if available, otherwise fall back to inline implementation."""
    if _handle_repairs_module:
        return _handle_repairs_module(whatsapp, phone_number, catalog_id=CATALOG_ID)

    # Fallback: load repair IDs from separate Excel file
    from catalog_utils import load_repair_retailer_ids
    repair_ids = load_repair_retailer_ids()
    if not repair_ids:
        repair_ids = _env_retailer_ids("PRODUCT_RETAILER_ID_REPAIR", "PRODUCT_RETAILER_ID_REPAIR_2")

    if not repair_ids:
        logger.warning("No repair retailer IDs configured (env or excel)")
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body="No repair packages are configured right now. Contact support to add products.",
            buttons=[ReplyButton(id="contact_support", title="Contact Support")],
        )
        return

    header = "SpectraX Repair Packages"
    body = "🛠 Choose a repair package. Includes diagnostics and software cleanup when registered."
    footer = "Tap a repair package to view details & book."

    try:
        # Use the safe catalog compatibility function
        from catalog_utils import send_catalog_compat
        send_catalog_compat(
            whatsapp=whatsapp,
            to=phone_number,
            retailer_ids=repair_ids,
            header=header,
            body=body,
            footer=footer,
            catalog_id=CATALOG_ID,
            fallback_button_id="browse_repairs"
        )
    except Exception as exc:
        logger.exception("Failed sending repair catalog: %s", exc)
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body="Sorry, something went wrong while fetching repair packages. Try again later.",
            buttons=[ReplyButton(id="try_again", title="Try Again")],
        )


def send_why_spectrax_message(phone_number: str):
    """Send why choose SpectraX Laptops message"""
    message = """✨ Why SpectraX Laptops?  
Because we don't just sell laptops — we provide a complete ecosystem for your digital success.  

✅ Premium laptop models with latest specs  
✅ FREE Starter Essentials software suite  
✅ Lifetime repair tracking & support  
✅ Real-time service updates via WhatsApp  
✅ Professional consultancy services  
✅ Trust & peace of mind in Zimbabwe  

🎯 **The Promise**: Your laptop + our expertise = unstoppable productivity!

Ready to explore? 👇"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="browse_laptops", title="💻 Browse Laptops"),
            ReplyButton(id="lifetime_support", title="🛡 Lifetime Support"),
        ],
    )


def send_lifetime_support_message(phone_number: str):
    """Send the lifetime support & benefits message with follow-up buttons"""
    message_body = """🎁 SpectraX Lifetime Support Package (Included with Registration)

**🔧 Lifetime Repair Services:**
• Dust cleaning & hardware maintenance
• Software troubleshooting & optimization
• Real-time repair tracking via WhatsApp
• Professional diagnostics & consultation

**💾 FREE Starter Essentials Software:**
• Microsoft Office alternatives
• Antivirus & security suite
• Photo/video editing tools
• Productivity & organization apps

**⚡ Premium Add-ons Available:**
• RAM/SSD upgrades during service
• Custom software installations
• Performance optimization packages
• Advanced business consultation

**📱 WhatsApp Integration:**
• Schedule services instantly
• Real-time repair updates
• Direct tech support chat
• Order tracking & notifications

✅ **Registration unlocks everything!** Simple, free, and gives you access to our complete ecosystem."""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message_body,
        buttons=[
            ReplyButton(id="browse_laptops", title="💻 Browse Laptops"),
            ReplyButton(id="register_laptop", title="� Register Laptop"),
            ReplyButton(id="schedule_service", title="🔧 Schedule Service"),
        ],
    )


def send_how_to_order_message(phone_number: str):
    """Send how to order information for laptops"""
    message_body = """How to Order Your SpectraX Laptop �

1️⃣ Tap "Browse Laptops" to view our catalog.

2️⃣ Click on any laptop model to see full specs.

3️⃣ Review features, RAM, storage, and pricing.

4️⃣ Tap "Add to Cart" for your chosen model.

5️⃣ Consider optional upgrades:
   • RAM/SSD upgrades
   • Premium software packages  
   • Accessories (bag, mouse, keyboard)
   • Extended warranty

6️⃣ Go to "View Cart" to review your order.

7️⃣ Adjust quantity and add-ons if needed.

8️⃣ Tap "Place Order" to confirm.

✅ You'll see "Order Successful", and we'll contact you shortly to finalize delivery, payment & schedule your FREE laptop registration!

🎁 **Don't forget**: Registration unlocks your FREE Starter Essentials software + lifetime repair tracking!

🎥 Want to see the ordering process?
Just tap below for a quick video demonstration."""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message_body,
        buttons=[
            ReplyButton(id="request_video_demo", title="🎥 Video Demo"),
            ReplyButton(id="browse_laptops", title="� Browse Laptops"),
        ],
    )


async def handle_video_demo_request(phone_number: str):
    """Handle video demo request for laptop ordering"""
    video_url = f"{PUBLIC_URL}/static/BUY_V1_Pro.mp4"
    
    # Send the video first
    whatsapp.send_video(
        to=phone_number,
        url=video_url,
        caption="🎥 Here's your SpectraX Laptop ordering demo!\n\nWatch how easy it is to browse laptops, select upgrades, and place your order through WhatsApp. �✨\n\nReady to get yours? Just tap 'Browse Laptops' below! 🛒",
    )
    
    # Wait 15 seconds before sending follow-up message
    await asyncio.sleep(15)
    
    # Send follow-up message with action buttons
    follow_up_message = """🎉 Thanks for watching the ordering demo! Getting your perfect laptop with lifetime support is super simple.

You can now click Browse Laptops below to explore our collection.
Don't forget - registration unlocks FREE software & lifetime repairs!

👇 Choose an option to continue:"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=follow_up_message,
        buttons=[
            ReplyButton(id="browse_laptops", title="💻 Browse Laptops"),
            ReplyButton(id="why_spectrax", title="💡 Why SpectraX?"),
            ReplyButton(id="lifetime_support", title="🛡 Lifetime Support"),
        ],
    )


def send_upgrades_accessories_message(phone_number: str):
    """Send upgrades and accessories options"""
    message_body = """⚡ Available Upgrades & Accessories

**🔧 Hardware Upgrades:**
• RAM upgrades (8GB → 16GB → 32GB)
• SSD storage upgrades (256GB → 512GB → 1TB)
• Performance optimization packages

**💾 Software Packages:**
• Premium productivity suites
• Creative software bundles
• Business & accounting tools
• Security & backup solutions

**🎒 Essential Accessories:**
• Premium laptop bags & cases
• Wireless mice & keyboards
• Cooling pads & stands
• External monitors & adapters

**🛡 Protection & Warranty:**
• Extended warranty (2-3 years)
• Accidental damage protection
• Insurance packages

💡 **Pro Tip**: Upgrades during purchase save you time and money!

Want to see the full laptop catalog with upgrade options?"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message_body,
        buttons=[
            ReplyButton(id="browse_laptops", title="💻 View Catalog"),
            ReplyButton(id="lifetime_support", title="🛡 Support Info"),
            ReplyButton(id="how_to_order", title="💳 How to Order"),
        ],
    )


def send_registration_flow(phone_number: str):
    """Send laptop registration information and process"""
    message_body = """📝 Laptop Registration - Unlock Your Benefits!

**Why Register Your Laptop?**
✅ Activates FREE Starter Essentials software suite
✅ Enables lifetime repair tracking
✅ Unlocks WhatsApp-based tech support
✅ Qualifies for priority service booking
✅ Access to exclusive upgrade offers

**What You Need:**
• Your laptop model & serial number
• Purchase receipt/order confirmation
• Your contact details

**Registration Process:**
1️⃣ Take a photo of your laptop's serial number sticker
2️⃣ Send it via WhatsApp with your order details
3️⃣ We'll confirm and activate your benefits
4️⃣ Receive your software download links
5️⃣ Start enjoying lifetime support!

**Already purchased a laptop?** 
Send us your order details and we'll help you register immediately!

**Haven't purchased yet?**
Browse our collection and registration will be included in your purchase process."""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message_body,
        buttons=[
            ReplyButton(id="browse_laptops", title="� Browse First"),
            ReplyButton(id="schedule_service", title="🔧 Need Service?"),
            ReplyButton(id="lifetime_support", title="🛡 Support Details"),
        ],
    )


def send_service_booking_flow(phone_number: str):
    """Send service booking options and process"""
    message_body = """🔧 Schedule Laptop Service & Support

**Available Services:**
• Software troubleshooting & optimization
• Hardware cleaning & maintenance  
• RAM/SSD upgrade installations
• Virus removal & security setup
• Custom software installations
• Performance diagnostics
• Business consultation services

**Service Types:**
🏠 **On-site Service** - We come to you
🏢 **Workshop Service** - Drop-off for detailed work
� **Remote Support** - Via WhatsApp/video call
📞 **Consultation** - Technical advice & planning

**How to Book:**
1️⃣ Describe your issue or service needed
2️⃣ Choose service type (on-site/workshop/remote)
3️⃣ Select preferred date & time
4️⃣ Confirm booking details
5️⃣ Receive confirmation & tracking info

**For Registered Laptops:** 
Many services are FREE or heavily discounted!

**Need help now?**
Just reply with your laptop issue and we'll guide you through the next steps."""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message_body,
        buttons=[
            ReplyButton(id="register_laptop", title="📝 Register First"),
            ReplyButton(id="browse_laptops", title="💻 Browse Laptops"),
            ReplyButton(id="lifetime_support", title="🛡 Support Info"),
        ],
    )


def send_admin_activity_stats(phone_number: str):
    """Send activity statistics to admin"""
    try:
        recent_activities = activity_logger.get_recent_activities(10)
        
        if not recent_activities:
            message = "📊 **Activity Statistics**\n\nNo recent activities found."
        else:
            message = "📊 **Recent Activity (Last 10)**\n\n"
            
            for activity in recent_activities:
                time_str = activity['timestamp']
                phone = activity['phone_number'][-4:]  # Last 4 digits for privacy
                activity_type = activity['activity_type']
                admin_flag = "👨‍💼" if activity['admin_flag'] else "👤"
                
                message += f"{admin_flag} {time_str} - ...{phone} - {activity_type}\n"
            
            # Get unique users count
            unique_users = len(set(activity['phone_number'] for activity in recent_activities))
            message += f"\n📈 **Summary:**\n"
            message += f"• Unique users: {unique_users}\n"
            message += f"• Total activities: {len(recent_activities)}\n"
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_back_main", title="⬅️ Back to Main"),
                ReplyButton(id="admin_order_management", title="📦 Orders"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to get activity stats")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading activity stats: {str(e)}")


def send_admin_analytics_menu(phone_number: str):
    """Send comprehensive analytics menu to admin"""
    try:
        # Get quick stats
        stats = activity_logger.get_analytics_summary(7)  # Last 7 days
        
        if "error" in stats:
            message = f"📊 **Analytics Dashboard**\n\n❌ {stats['error']}"
        else:
            message = f"""📊 **Analytics Dashboard (Last 7 Days)**

**📈 Quick Overview:**
• Total Activities: {stats['total_activities']}
• Unique Users: {stats['unique_users']}
• Admin Actions: {stats['admin_activities']}
• User Actions: {stats['user_activities']}
• Avg per User: {stats['avg_activities_per_user']}

**🔥 Top Activities:**"""
            
            for activity, count in stats['top_activity_types'][:3]:
                message += f"\n• {activity}: {count}"
            
            message += f"\n\n**⏰ Peak Hours:**"
            for hour, count in stats['peak_hours'][:2]:
                time_period = "AM" if int(hour) < 12 else "PM"
                display_hour = int(hour) if int(hour) <= 12 else int(hour) - 12
                if display_hour == 0:
                    display_hour = 12
                message += f"\n• {display_hour}:00 {time_period}: {count} activities"
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_detailed_analytics", title="📊 Details"),
                ReplyButton(id="admin_export_data", title="📥 Export"),
                ReplyButton(id="admin_conversation_analytics", title="💬 Conversations"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to get analytics menu")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading analytics: {str(e)}")


def send_admin_detailed_analytics(phone_number: str):
    """Send detailed analytics breakdown"""
    try:
        stats_7d = activity_logger.get_analytics_summary(7)
        stats_30d = activity_logger.get_analytics_summary(30)
        conv_stats = activity_logger.get_conversation_analytics()
        
        message = """📊 **Detailed Analytics Report**

**📅 7-Day vs 30-Day Comparison:**"""
        
        if "error" not in stats_7d and "error" not in stats_30d:
            message += f"""
• Users (7d/30d): {stats_7d['unique_users']} / {stats_30d['unique_users']}
• Activities (7d/30d): {stats_7d['total_activities']} / {stats_30d['total_activities']}
• Sessions (7d/30d): {stats_7d['total_sessions']} / {stats_30d['total_sessions']}

**💬 Conversation Insights:**"""
            
            if "error" not in conv_stats:
                message += f"""
• Total Conversations: {conv_stats['total_conversations']}
• Avg Duration: {conv_stats['avg_conversation_duration_minutes']} min
• Longest Chat: {conv_stats['longest_conversation_minutes']} min
• Most Engaged Users: {len([u for u, data in conv_stats['user_engagement'].items() if data['total_activities'] > 5])}

**🎯 User Engagement Levels:**
• High (10+ activities): {len([u for u, data in conv_stats['user_engagement'].items() if data['total_activities'] >= 10])} users
• Medium (5-9 activities): {len([u for u, data in conv_stats['user_engagement'].items() if 5 <= data['total_activities'] < 10])} users  
• Low (1-4 activities): {len([u for u, data in conv_stats['user_engagement'].items() if data['total_activities'] < 5])} users"""
            
            # Daily activity trend
            if stats_7d['daily_breakdown']:
                sorted_days = sorted(stats_7d['daily_breakdown'].items())
                message += f"\n\n**📅 Daily Activity (Last 7 Days):**"
                for day, count in sorted_days[-7:]:
                    date_obj = datetime.strptime(day, "%Y-%m-%d")
                    day_name = date_obj.strftime("%a")
                    message += f"\n• {day_name} {day}: {count} activities"
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_export_detailed", title="📥 Export Report"),
                ReplyButton(id="admin_analytics_menu", title="⬅️ Analytics"),
                ReplyButton(id="admin_back_main", title="🏠 Main"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to get detailed analytics")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading detailed analytics: {str(e)}")


def send_admin_conversation_analytics(phone_number: str):
    """Send conversation-focused analytics"""
    try:
        conv_stats = activity_logger.get_conversation_analytics()
        
        if "error" in conv_stats:
            message = f"💬 **Conversation Analytics**\n\n❌ {conv_stats['error']}"
        else:
            message = f"""💬 **Conversation Analytics**

**📊 Overview:**
• Total Conversations: {conv_stats['total_conversations']}
• Unique Users: {conv_stats['total_users']}
• Avg Duration: {conv_stats['avg_conversation_duration_minutes']} minutes

**⏱️ Duration Insights:**
• Longest Chat: {conv_stats['longest_conversation_minutes']} min
• Shortest Chat: {conv_stats['shortest_conversation_minutes']} min

**🏆 Top Engaged Users:**"""
            
            # Show top 5 most engaged users
            top_users = sorted(
                conv_stats['user_engagement'].items(), 
                key=lambda x: x[1]['total_activities'], 
                reverse=True
            )[:5]
            
            for i, (phone, data) in enumerate(top_users, 1):
                masked_phone = "..." + phone[-4:] if len(phone) > 4 else phone
                message += f"\n{i}. {masked_phone}: {data['total_activities']} activities ({data['session_count']} sessions)"
            
            message += f"\n\n**📈 User Behavior:**"
            # Analyze common activity patterns
            all_top_activities = [data['top_activity'] for data in conv_stats['user_engagement'].values()]
            activity_frequency = {}
            for activity in all_top_activities:
                activity_frequency[activity] = activity_frequency.get(activity, 0) + 1
            
            top_behaviors = sorted(activity_frequency.items(), key=lambda x: x[1], reverse=True)[:3]
            for activity, count in top_behaviors:
                message += f"\n• {activity}: {count} users prefer this"
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_user_details", title="👤 Users"),
                ReplyButton(id="admin_analytics_menu", title="⬅️ Analytics"),
                ReplyButton(id="admin_export_conversations", title="📥 Export"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to get conversation analytics")
        whatsapp.send_text(to=phone_number, body=f"❌ Error loading conversation analytics: {str(e)}")


def send_admin_export_menu(phone_number: str):
    """Send data export options menu"""
    message = """📥 **Data Export Options**

**📊 Available Exports:**

**📋 Quick Exports:**
• Last 7 days (all data)
• Last 30 days (all data)  
• Admin activities only
• User activities only
• Recent conversations

**🎯 Custom Exports:**
• Specific date range
• Filtered by activity type
• User engagement report
• Conversation analysis

**📁 Export Formats:**
All exports are provided as Excel files (.xlsx) with formatted data and summary sheets.

**⚡ Quick Actions:**"""
    
    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_export_7days", title="📅 7 Days"),
            ReplyButton(id="admin_export_30days", title="📅 30 Days"),
            ReplyButton(id="admin_export_admin_only", title="👨‍💼 Admin Only"),
        ],
    )


def handle_admin_export_request(phone_number: str, export_type: str):
    """Handle different types of export requests"""
    try:
        from datetime import datetime, timedelta
        
        export_file = f"spectrax_export_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        success = False
        
        if export_type == "7days":
            start_date = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
            success = activity_logger.export_filtered_data(
                start_date=start_date,
                output_file=export_file
            )
            description = "Last 7 days activity"
            
        elif export_type == "30days":
            start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            success = activity_logger.export_filtered_data(
                start_date=start_date,
                output_file=export_file
            )
            description = "Last 30 days activity"
            
        elif export_type == "admin_only":
            success = activity_logger.export_filtered_data(
                admin_only=True,
                output_file=export_file
            )
            description = "Admin activities only"
            
        elif export_type == "conversations":
            # Export conversation-focused data
            success = activity_logger.export_filtered_data(
                activity_types=["message_received", "button_clicked", "order_placed"],
                output_file=export_file
            )
            description = "Conversation activities"
        
        if success:
            # Get file size for info
            file_size = os.path.getsize(export_file) if os.path.exists(export_file) else 0
            file_size_mb = round(file_size / (1024 * 1024), 2)
            
            message = f"""✅ **Export Complete!**

**📁 File Generated:**
• Name: {export_file}
• Content: {description}
• Size: {file_size_mb} MB
• Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

**📥 Download Instructions:**
The file is ready in your project directory. You can access it via:
• File manager on server
• Download via admin panel
• FTP/SFTP transfer

**📊 Next Steps:**
• Open in Excel for analysis
• Share with team members
• Create reports and insights"""
            
        else:
            message = f"❌ **Export Failed**\n\nFailed to generate {description} export. Please check the logs and try again."
        
        whatsapp.send_interactive_buttons(
            to=phone_number,
            body=message,
            buttons=[
                ReplyButton(id="admin_export_menu", title="📥 Exports"),
                ReplyButton(id="admin_analytics_menu", title="📊 Analytics"),
                ReplyButton(id="admin_back_main", title="🏠 Main"),
            ],
        )
        
    except Exception as e:
        logger.exception("Failed to handle export request")
        whatsapp.send_text(to=phone_number, body=f"❌ Export error: {str(e)}")


def send_admin_order_processing_menu(phone_number: str):
    """Send order processing options to admin"""
    message = """⚡ **Order Processing Center**

**📋 Quick Actions:**
• Mark order as processing
• Contact customer for details
• Request payment confirmation
• Schedule delivery/pickup
• Update order status

**📞 Customer Communication:**
• Send order confirmation
• Request additional details
• Provide delivery updates
• Handle special requests

**📦 Fulfillment Options:**
• Prepare items for shipment
• Schedule installation
• Arrange pickup service
• Process returns/exchanges"""

    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_mark_processing", title="🔄 Processing"),
            ReplyButton(id="admin_request_payment", title="💳 Payment"),
            ReplyButton(id="admin_schedule_delivery", title="🚚 Delivery"),
        ],
    )


def send_admin_contact_customer_menu(phone_number: str):
    """Send customer contact options to admin"""
    message = """📞 **Customer Contact Center**

**🎯 Contact Purposes:**
• Order confirmation & details
• Payment method confirmation
• Delivery scheduling
• Technical specifications
• Special requests/customizations

**💬 Communication Templates:**
• Order received confirmation
• Payment request message
• Delivery scheduling
• Technical support follow-up
• Thank you & feedback request

**📋 Customer Information:**
Use recent order details to personalize communication and provide excellent service."""

    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_send_confirmation", title="✅ Confirm"),
            ReplyButton(id="admin_request_details", title="📝 Details"),
            ReplyButton(id="admin_schedule_call", title="📞 Call"),
        ],
    )


def send_admin_order_details_menu(phone_number: str):
    """Send order details and management options"""
    message = """📋 **Order Details & Management**

**📊 Order Analytics:**
• View complete order history
• Customer purchase patterns
• Product performance metrics
• Revenue tracking

**🔄 Order Management:**
• Update order status
• Modify order details
• Process cancellations
• Handle returns/exchanges

**📈 Business Intelligence:**
• Sales performance
• Customer satisfaction
• Inventory insights
• Growth opportunities"""

    whatsapp.send_interactive_buttons(
        to=phone_number,
        body=message,
        buttons=[
            ReplyButton(id="admin_view_all_orders", title="📊 All Orders"),
            ReplyButton(id="admin_update_status", title="🔄 Update"),
            ReplyButton(id="admin_back_main", title="🏠 Main"),
        ],
    )


def _get_text_content(msg) -> Optional[str]:
    """Safely extract text content from a message."""
    try:
        if hasattr(msg, 'text'):
            return msg.text
        elif hasattr(msg, 'body'):
            return msg.body
        return None
    except Exception:
        return None
    """Best-effort extraction of text content from a TextMessage.
    Tries common fields: msg.text (str or object with .body), msg.body, dict-like access.
    Returns None if not found.
    """
    try:
        # direct string
        val = getattr(msg, "text", None)
        if isinstance(val, str) and val.strip():
            return val.strip()
        # object with body attribute
        if hasattr(val, "body"):
            b = getattr(val, "body")
            if isinstance(b, str) and b.strip():
                return b.strip()
        # some SDKs expose .body directly
        body = getattr(msg, "body", None)
        if isinstance(body, str) and body.strip():
            return body.strip()
        # dict-like text
        if isinstance(val, dict):
            b2 = val.get("body")
            if isinstance(b2, str) and b2.strip():
                return b2.strip()
    except Exception:
        pass
    return None


if __name__ == "__main__":
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
