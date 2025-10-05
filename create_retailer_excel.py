"""
create_retailer_excel.py

Creates an Excel workbook with two sheets:
- "Laptops" containing PRODUCT_RETAILER_ID and PRODUCT_RETAILER_ID_2 (if set)
- "Repairs" containing PRODUCT_RETAILER_ID_REPAIR and PRODUCT_RETAILER_ID_REPAIR_2 (if set)

Writes file: spectrax_retailer_ids.xlsx
"""
import os
from typing import List
from dotenv import load_dotenv
from openpyxl import Workbook

load_dotenv()


def _collect_retailer_ids(*keys: str) -> List[str]:
    """Return non-empty environment values for the given keys."""
    ids = []
    for k in keys:
        val = os.getenv(k)
        if val and val.strip():
            ids.append(val.strip())
    return ids


def _write_sheet(workbook: Workbook, title: str, header: str, values: List[str]) -> None:
    ws = workbook.create_sheet(title=title) if title not in workbook.sheetnames else workbook[title]
    # Clear existing content if any
    ws.delete_rows(1, ws.max_row)
    ws.append([header])
    if not values:
        ws.append(["(none configured)"])
    else:
        for v in values:
            ws.append([v])


def main() -> None:
    laptop_keys = ("PRODUCT_RETAILER_ID", "PRODUCT_RETAILER_ID_2")
    repair_keys = ("PRODUCT_RETAILER_ID_REPAIR", "PRODUCT_RETAILER_ID_REPAIR_2")

    laptop_ids = _collect_retailer_ids(*laptop_keys)
    repair_ids = _collect_retailer_ids(*repair_keys)

    wb = Workbook()
    # Replace default sheet with Laptops sheet
    default_title = wb.active.title
    wb.active.title = "Laptops"
    _write_sheet(wb, "Laptops", "retailer_id", laptop_ids)
    _write_sheet(wb, "Repairs", "retailer_id", repair_ids)

    # Remove any extra default sheet if present and unnamed
    if default_title not in ("Laptops", "Repairs") and default_title in wb.sheetnames and default_title != wb.active.title:
        try:
            wb.remove(wb[default_title])
        except Exception:
            pass

    out_filename = "spectrax_retailer_ids.xlsx"
    wb.save(out_filename)
    print(f"Wrote retailer IDs to {out_filename}")


if __name__ == "__main__":
    main()